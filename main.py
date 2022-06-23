import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import os
import datetime


def scrapCard(card):

    try: title = card.find('h3').find('a').text
    except: title = 'x'

    try: link = card.find('h3').find('a')['href']
    except: link = "x"
    full_link = 'https://zewailcity.edu.eg/main/' + link

    try: time = card.time.em.strong.text
    except: time = "x"

    try: paragraph = card.find('div', class_='news-text').p.text
    except:
        try: paragraph = card.find('div', class_='news-text').div.text
        except:paragraph = 'x'

    try: photo = card.img['src'].replace('../', 'https://www.zewailcity.edu.eg/')
    except: photo = 'x'

    return title, full_link, time, paragraph, photo


def savePageToSheet(pageCards, ws):
    for card in pageCards:
        title, full_link, time, paragraph, photo_link = scrapCard(card)

        row = ws.max_row+1
        ws.cell(row=row, column=1).value = title
        ws.cell(row=row, column=2).value = time
        ws.cell(row=row, column=3).value = 'link'
        ws.cell(row=row, column=3).hyperlink = full_link
        ws.cell(row=row, column=3).style = "Hyperlink"
        ws.cell(row=row, column=4).value = 'photo'
        ws.cell(row=row, column=4).hyperlink = photo_link
        ws.cell(row=row, column=4).style = "Hyperlink"
        ws.cell(row=row, column=5).value = paragraph


def scrapSite(ws):

    response = requests.get(f'https://www.zewailcity.edu.eg/main/content.php?lang=en&alias=recent_news')
    lastPageCard = BeautifulSoup(response.text, 'lxml').find('div', class_='page_num').ul.find_all('li')[-2]
    lastPageNum = int(lastPageCard.text)


    for page in range(1,2):
        response = requests.get(f'https://www.zewailcity.edu.eg/main/content.php?lang=en&alias=recent_news&page={page}')

        soap = BeautifulSoup(response.text,'lxml')
        pageCards = soap.find_all('div', class_= 'news-content clearfix')

        savePageToSheet(pageCards, ws)


def addExcelHeader(ws):
    today = datetime.date.today()
    ws['A1'] = f'This excel contains data scrapped from Zewail City news from website creation and until {today}.'
    ws.cell(row=2, column=1).value = 'title'
    ws.cell(row=2, column=2).value = 'time'
    ws.cell(row=2, column=3).value = 'full_link'
    ws.cell(row=2, column=4).value = 'photo_link'
    ws.cell(row=2, column=5).value = 'paragraph'


def excelStyling(ws):
    ws.column_dimensions['A'].width =80
    ws.column_dimensions['B'].width = 19
    ws.column_dimensions['C'].width = 8.11 #default
    ws.column_dimensions['D'].width = 8.11 #default
    ws.column_dimensions['E'].width = 80



def main():
    wb = Workbook()
    ws = wb.active

    addExcelHeader(ws)
    scrapSite(ws)
    excelStyling(ws)


    try:
        if not os.path.exists('./output'):
            os.makedirs('./output')
        wb.save('./output/zc-news.xlsx')
        print(f'saved in ./output/zc-news.xlsx')
    except:
        i = 0
        while True:
            try:
                i += 1
                wb.save(f'./output/zc-news-{i}.xlsx')
                print(f'saved in ./output/zc-news-{i}.xlsx')
                return
            except:
                pass

if __name__ == '__main__':
    main()